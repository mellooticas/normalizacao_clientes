"""
Análise detalhada da estrutura de vendas
Identificando o limite inferior "Restante Entrada"
"""

import openpyxl
from pathlib import Path

# Arquivo de exemplo
arquivo = Path('dados_processados/originais/cxs/planilhas_originais/maua/jan_24.xlsx')

print('=' * 80)
print('ANÁLISE COMPLETA - ESTRUTURA DE VENDAS COM LIMITE')
print('=' * 80)
print(f'\nArquivo: {arquivo.name}')

# Carregar workbook
wb = openpyxl.load_workbook(arquivo, data_only=True)

# Analisar aba '01'
print(f'\n{"=" * 80}')
print('ABA: 01 (DIA 1)')
print('=' * 80)

ws = wb['01']

# 1. Procurar "Restante Entrada"
print('\n🔍 BUSCANDO "RESTANTE ENTRADA":')

linha_restante = None
for row_idx in range(1, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, col_idx)
        if cell.value and 'restante entrada' in str(cell.value).lower():
            linha_restante = row_idx
            letra = openpyxl.utils.get_column_letter(col_idx)
            print(f'\n✅ Encontrado em {letra}{row_idx}: "{cell.value}"')
            break
    if linha_restante:
        break

# 2. Mostrar estrutura completa da coluna E até I
print(f'\n{"=" * 80}')
print('ESTRUTURA COMPLETA - COLUNAS E, F, G, H, I')
print('=' * 80)

colunas_vendas = {
    'E': 'Nº Venda',
    'F': 'Cliente', 
    'G': 'Forma de Pgto',
    'H': 'Valor Venda',
    'I': 'Entrada'
}

# Mostrar linha por linha desde linha 1 até "Restante Entrada"
limite = linha_restante if linha_restante else 50

print(f'\n📋 Linhas 1 até {limite} (onde está "Restante Entrada"):')
print('=' * 80)

for row_idx in range(1, limite + 5):  # +5 para ver um pouco depois também
    valores = {}
    tem_conteudo = False
    
    for letra, nome in colunas_vendas.items():
        col_idx = openpyxl.utils.column_index_from_string(letra)
        cell = ws.cell(row_idx, col_idx)
        
        if cell.value is not None:
            valores[letra] = str(cell.value)[:40]  # Limitar tamanho
            tem_conteudo = True
        else:
            valores[letra] = ''
    
    if tem_conteudo:
        # Formatar linha
        linha_str = f"Linha {row_idx:3d}: "
        
        for letra in ['E', 'F', 'G', 'H', 'I']:
            if valores[letra]:
                linha_str += f"{letra}={valores[letra][:20]:20s} | "
        
        print(linha_str)
        
        # Destacar linha do cabeçalho
        if 'Nº Venda' in valores['E']:
            print('   ↑ CABEÇALHO')
        
        # Destacar linha "Restante Entrada"
        if row_idx == linha_restante:
            print('   ↓ LIMITE INFERIOR - RESTANTE ENTRADA')

# 3. Contar vendas válidas
print(f'\n{"=" * 80}')
print('CONTAGEM DE VENDAS VÁLIDAS')
print('=' * 80)

vendas_validas = []
linha_inicio = 6  # Linha 6 é onde começam os dados (após linha 5 de cabeçalho)
linha_fim = linha_restante - 1 if linha_restante else 40

print(f'\n📊 Analisando linhas {linha_inicio} a {linha_fim}')

for row_idx in range(linha_inicio, linha_fim + 1):
    col_e = ws.cell(row_idx, openpyxl.utils.column_index_from_string('E'))
    col_f = ws.cell(row_idx, openpyxl.utils.column_index_from_string('F'))
    
    # Verificar se tem número de venda E cliente
    if col_e.value and col_f.value:
        # Verificar se o número de venda é um número
        try:
            num_venda = str(col_e.value).strip()
            if num_venda and not num_venda.lower() in ['restante entrada', 'nº venda', 'vendas']:
                vendas_validas.append({
                    'linha': row_idx,
                    'num_venda': num_venda,
                    'cliente': str(col_f.value)[:30]
                })
        except:
            pass

print(f'\n✅ Total de vendas válidas encontradas: {len(vendas_validas)}')

if vendas_validas:
    print(f'\n📋 Primeiras 5 vendas:')
    for i, venda in enumerate(vendas_validas[:5], 1):
        print(f"   {i}. Linha {venda['linha']}: OS {venda['num_venda']} - {venda['cliente']}")
    
    print(f'\n📋 Últimas 5 vendas:')
    for i, venda in enumerate(vendas_validas[-5:], 1):
        print(f"   {i}. Linha {venda['linha']}: OS {venda['num_venda']} - {venda['cliente']}")

# 4. Analisar outra aba para confirmar padrão
print(f'\n{"=" * 80}')
print('ABA: 15 (DIA 15 - PARA COMPARAÇÃO)')
print('=' * 80)

if '15' in wb.sheetnames:
    ws15 = wb['15']
    
    # Procurar "Restante Entrada"
    linha_restante_15 = None
    for row_idx in range(1, ws15.max_row + 1):
        for col_idx in range(1, ws15.max_column + 1):
            cell = ws15.cell(row_idx, col_idx)
            if cell.value and 'restante entrada' in str(cell.value).lower():
                linha_restante_15 = row_idx
                print(f'\n✅ "Restante Entrada" encontrado na linha {row_idx}')
                break
        if linha_restante_15:
            break
    
    # Contar vendas
    vendas_dia15 = []
    for row_idx in range(6, (linha_restante_15 if linha_restante_15 else 40)):
        col_e = ws15.cell(row_idx, openpyxl.utils.column_index_from_string('E'))
        col_f = ws15.cell(row_idx, openpyxl.utils.column_index_from_string('F'))
        
        if col_e.value and col_f.value:
            try:
                num_venda = str(col_e.value).strip()
                if num_venda and not num_venda.lower() in ['restante entrada', 'nº venda', 'vendas']:
                    vendas_dia15.append(num_venda)
            except:
                pass
    
    print(f'✅ Total de vendas no dia 15: {len(vendas_dia15)}')

wb.close()

print(f'\n{"=" * 80}')
print('✅ Análise concluída!')
print('=' * 80)
