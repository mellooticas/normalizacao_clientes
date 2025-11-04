#!/usr/bin/env python3
"""
Script para analisar arquivos Excel de CAIXA e identificar abas e tabelas nomeadas
Sistema Carne Fácil - Análise de dados de caixa
"""

import pandas as pd
import openpyxl
from pathlib import Path
import json
from datetime import datetime
import os
import warnings
warnings.filterwarnings('ignore')

def analisar_arquivo_excel(caminho_arquivo):
    """
    Analisa um arquivo Excel e retorna informações sobre abas e tabelas nomeadas
    """
    resultado = {
        'arquivo': str(caminho_arquivo),
        'tamanho_kb': round(os.path.getsize(caminho_arquivo) / 1024, 2),
        'abas': {},
        'tabelas_nomeadas': [],
        'erro': None
    }
    
    try:
        # Carregar workbook com openpyxl para acessar tabelas nomeadas
        workbook = openpyxl.load_workbook(caminho_arquivo, data_only=True)
        
        # Listar todas as abas
        for nome_aba in workbook.sheetnames:
            try:
                # Análise básica da aba
                worksheet = workbook[nome_aba]
                
                # Contar células com dados
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                # Tentar carregar com pandas para mais análises
                try:
                    df = pd.read_excel(caminho_arquivo, sheet_name=nome_aba, header=None)
                    linhas_com_dados = len(df.dropna(how='all'))
                    colunas_com_dados = len(df.dropna(how='all', axis=1).columns)
                except Exception:
                    linhas_com_dados = max_row
                    colunas_com_dados = max_col
                
                # Verificar se há cabeçalhos na primeira linha
                primeira_linha = []
                try:
                    for col in range(1, min(max_col + 1, 21)):  # Máximo 20 colunas
                        valor = worksheet.cell(row=1, column=col).value
                        if valor:
                            primeira_linha.append(str(valor).strip())
                        else:
                            primeira_linha.append('')
                except Exception:
                    pass
                
                resultado['abas'][nome_aba] = {
                    'max_linha': max_row,
                    'max_coluna': max_col,
                    'linhas_com_dados': linhas_com_dados,
                    'colunas_com_dados': colunas_com_dados,
                    'primeira_linha': primeira_linha[:10],  # Primeiras 10 colunas
                    'tem_dados': linhas_com_dados > 1
                }
                
            except Exception as e:
                resultado['abas'][nome_aba] = {
                    'erro': str(e)
                }
        
        # Tentar identificar tabelas nomeadas
        try:
            # Tabelas nomeadas (defined names)
            for nome_definido in workbook.defined_names:
                try:
                    if nome_definido.name and '!' in str(nome_definido.value):
                        resultado['tabelas_nomeadas'].append({
                            'nome': nome_definido.name,
                            'referencia': str(nome_definido.value),
                            'tipo': 'named_range'
                        })
                except Exception:
                    pass
            
            # Tabelas estruturadas (Tables)
            for nome_aba in workbook.sheetnames:
                worksheet = workbook[nome_aba]
                if hasattr(worksheet, '_tables'):
                    for tabela in worksheet._tables:
                        try:
                            resultado['tabelas_nomeadas'].append({
                                'nome': tabela.name if hasattr(tabela, 'name') else f'Table_{nome_aba}',
                                'aba': nome_aba,
                                'referencia': str(tabela.ref) if hasattr(tabela, 'ref') else 'N/A',
                                'tipo': 'table'
                            })
                        except Exception:
                            pass
                            
        except Exception as e:
            resultado['erro_tabelas'] = str(e)
        
        workbook.close()
        
    except Exception as e:
        resultado['erro'] = str(e)
    
    return resultado

def explorar_pasta_caixa():
    """
    Explora todas as pastas CAIXA das lojas e analisa os arquivos Excel
    """
    
    # Caminhos das pastas CAIXA
    caminhos_base = [
        "D:/OneDrive - Óticas Taty Mello/LOJAS/MAUA/CAIXA",
        "D:/OneDrive - Óticas Taty Mello/LOJAS/PERUS/CAIXA", 
        "D:/OneDrive - Óticas Taty Mello/LOJAS/RIO_PEQUENO/CAIXA",
        "D:/OneDrive - Óticas Taty Mello/LOJAS/SAO_MATEUS/CAIXA",
        "D:/OneDrive - Óticas Taty Mello/LOJAS/SUZANO/CAIXA",
        "D:/OneDrive - Óticas Taty Mello/LOJAS/SUZANO2/CAIXA"
    ]
    
    resultado_geral = {
        'data_analise': datetime.now().isoformat(),
        'lojas': {},
        'resumo': {
            'total_arquivos': 0,
            'abas_unicas': set(),
            'tabelas_nomeadas_unicas': set(),
            'padroes_identificados': {}
        }
    }
    
    print("🔍 Iniciando análise dos arquivos de CAIXA...")
    
    for caminho_loja in caminhos_base:
        nome_loja = caminho_loja.split('/')[-2]  # Extrair nome da loja
        
        if not os.path.exists(caminho_loja):
            print(f"⚠️  Pasta não encontrada: {caminho_loja}")
            continue
            
        print(f"\n📁 Analisando loja: {nome_loja}")
        
        resultado_geral['lojas'][nome_loja] = {
            'caminho': caminho_loja,
            'arquivos': {},
            'total_arquivos': 0,
            'abas_encontradas': set(),
            'tabelas_encontradas': []
        }
        
        # Buscar todos os arquivos Excel recursivamente
        arquivos_excel = []
        for root, dirs, files in os.walk(caminho_loja):
            for file in files:
                if file.endswith('.xlsx') and not file.startswith('~$'):
                    arquivos_excel.append(os.path.join(root, file))
        
        print(f"   📊 Encontrados {len(arquivos_excel)} arquivos Excel")
        
        # Analisar cada arquivo
        for i, caminho_arquivo in enumerate(arquivos_excel[:5], 1):  # Limitar a 5 por loja inicialmente
            nome_arquivo = os.path.basename(caminho_arquivo)
            print(f"   📄 Analisando {i}/5: {nome_arquivo}")
            
            resultado_arquivo = analisar_arquivo_excel(caminho_arquivo)
            resultado_geral['lojas'][nome_loja]['arquivos'][nome_arquivo] = resultado_arquivo
            
            # Coletar estatísticas
            if 'abas' in resultado_arquivo:
                for nome_aba in resultado_arquivo['abas'].keys():
                    resultado_geral['lojas'][nome_loja]['abas_encontradas'].add(nome_aba)
                    resultado_geral['resumo']['abas_unicas'].add(nome_aba)
            
            if 'tabelas_nomeadas' in resultado_arquivo:
                for tabela in resultado_arquivo['tabelas_nomeadas']:
                    resultado_geral['lojas'][nome_loja]['tabelas_encontradas'].append(tabela)
                    resultado_geral['resumo']['tabelas_nomeadas_unicas'].add(tabela['nome'])
            
            resultado_geral['resumo']['total_arquivos'] += 1
        
        resultado_geral['lojas'][nome_loja]['total_arquivos'] = len(arquivos_excel)
        resultado_geral['lojas'][nome_loja]['abas_encontradas'] = list(resultado_geral['lojas'][nome_loja]['abas_encontradas'])
    
    # Converter sets para listas para JSON
    resultado_geral['resumo']['abas_unicas'] = sorted(list(resultado_geral['resumo']['abas_unicas']))
    resultado_geral['resumo']['tabelas_nomeadas_unicas'] = sorted(list(resultado_geral['resumo']['tabelas_nomeadas_unicas']))
    
    return resultado_geral

def main():
    """Função principal"""
    print("🚀 ANÁLISE DE ARQUIVOS DE CAIXA - SISTEMA CARNE FÁCIL")
    print("=" * 60)
    
    # Executar análise
    resultado = explorar_pasta_caixa()
    
    # Salvar resultado completo
    caminho_resultado = 'data/originais/cxs/analise_estrutura_caixa.json'
    os.makedirs(os.path.dirname(caminho_resultado), exist_ok=True)
    
    with open(caminho_resultado, 'w', encoding='utf-8') as f:
        json.dump(resultado, f, indent=2, ensure_ascii=False)
    
    print(f"\n💾 Resultado completo salvo em: {caminho_resultado}")
    
    # Mostrar resumo
    print("\n📊 RESUMO DA ANÁLISE:")
    print(f"   📁 Total de arquivos analisados: {resultado['resumo']['total_arquivos']}")
    print(f"   📄 Abas únicas encontradas: {len(resultado['resumo']['abas_unicas'])}")
    print(f"   🏷️  Tabelas nomeadas únicas: {len(resultado['resumo']['tabelas_nomeadas_unicas'])}")
    
    print("\n📋 ABAS ENCONTRADAS:")
    for aba in resultado['resumo']['abas_unicas']:
        print(f"   • {aba}")
    
    if resultado['resumo']['tabelas_nomeadas_unicas']:
        print("\n🏷️ TABELAS NOMEADAS ENCONTRADAS:")
        for tabela in resultado['resumo']['tabelas_nomeadas_unicas']:
            print(f"   • {tabela}")
    
    print("\n📈 ANÁLISE POR LOJA:")
    for nome_loja, dados_loja in resultado['lojas'].items():
        print(f"\n   🏪 {nome_loja}:")
        print(f"      📊 Total de arquivos Excel: {dados_loja['total_arquivos']}")
        print(f"      📄 Abas encontradas: {len(dados_loja['abas_encontradas'])}")
        if dados_loja['abas_encontradas']:
            print(f"         {', '.join(dados_loja['abas_encontradas'])}")
        print(f"      🏷️  Tabelas: {len(dados_loja['tabelas_encontradas'])}")
    
    print("\n✅ Análise concluída!")
    return resultado

if __name__ == "__main__":
    main()