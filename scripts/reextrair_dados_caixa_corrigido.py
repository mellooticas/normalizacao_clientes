#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema CORRIGIDO de extração de dados do caixa
Corrige problemas de headers sendo extraídos como dados
"""

import pandas as pd
import openpyxl
import os
import json
from datetime import datetime
import re

# Estruturas corrigidas com validação de headers
ESTRUTURAS_TABELAS_CORRIGIDAS = {
    'VENDAS': {
        'celula_base': 'E5',
        'colunas_esperadas': ['Nº Venda', 'Cliente', 'Forma de Pgto', 'Valor Venda', 'Entrada'],
        'colunas_normalizadas': ['nn_venda', 'cliente', 'forma_de_pgto', 'valor_venda', 'entrada'],
        'validacao_numerica': ['nn_venda', 'valor_venda', 'entrada'],
        'tipo': 'financeiro'
    },
    'RESTANTE_ENTRADA': {
        'celula_base': 'E25',
        'colunas_esperadas': ['Nº Venda', 'Cliente', 'Forma de Pgto', 'Valor Venda', 'Entrada'],
        'colunas_normalizadas': ['nn_venda', 'cliente', 'forma_de_pgto', 'valor_venda', 'entrada'],
        'validacao_numerica': ['nn_venda', 'valor_venda', 'entrada'],
        'tipo': 'financeiro'
    },
    'RECEBIMENTO_CARNE': {
        'celula_base': 'E34',
        'colunas_esperadas': ['OS', 'Cliente', 'Forma de Pgto', 'Valor Parcela', 'Nº Parcela'],
        'colunas_normalizadas': ['os', 'cliente', 'forma_de_pgto', 'valor_parcela', 'n_parcela'],
        'validacao_numerica': ['valor_parcela', 'n_parcela'],
        'tipo': 'recebimento'
    },
    'OS_ENTREGUES_DIA': {
        'celula_base': 'K14',
        'colunas_esperadas': ['OS', 'Vendedor', 'CARNE'],
        'colunas_normalizadas': ['os', 'vendedor', 'carne'],
        'validacao_numerica': [],
        'tipo': 'operacional'
    },
    'ENTREGA_CARNE': {
        'celula_base': 'K34',
        'colunas_esperadas': ['OS', 'Parcelas', 'Valor Total'],
        'colunas_normalizadas': ['os', 'parcelas', 'valor_total'],
        'validacao_numerica': ['parcelas', 'valor_total'],
        'tipo': 'entrega'
    }
}

def converter_coordenada_para_indices(coordenada):
    """Converte coordenada Excel para índices"""
    from openpyxl.utils import coordinate_to_tuple
    return coordinate_to_tuple(coordenada)

def validar_se_eh_header(valor, colunas_esperadas):
    """Valida se um valor é um header da planilha"""
    if not valor:
        return False
    
    valor_str = str(valor).strip()
    
    # Lista de headers conhecidos que devem ser rejeitados
    headers_conhecidos = [
        'OS', 'Cliente', 'Forma de Pgto', 'Valor Venda', 'Entrada',
        'Valor Parcela', 'Nº Parcela', 'Nº Venda', 'Vendedor', 'CARNE',
        'Parcelas', 'Valor Total', 'Coluna1', 'Coluna2', 'Coluna3', 'Coluna4', 'Coluna5'
    ]
    
    # Verificar se é exatamente um header conhecido
    if valor_str in headers_conhecidos:
        return True
    
    # Verificar se é uma lista de headers (formato ['header1', 'header2'])
    if valor_str.startswith('[') and valor_str.endswith(']'):
        return True
    
    # Verificar se contém apenas texto de header sem números ou dados válidos
    if valor_str in colunas_esperadas:
        return True
    
    return False

def validar_linha_dados(linha_dados, estrutura):
    """Valida se uma linha contém dados válidos ou apenas headers"""
    
    # Verificar se algum valor é claramente um header
    for coluna_norm, valor in linha_dados.items():
        if coluna_norm in ['linha_origem', 'tabela_tipo', 'aba_origem', 'valores_brutos']:
            continue
            
        if validar_se_eh_header(valor, estrutura['colunas_esperadas']):
            return False
    
    # Verificar se pelo menos uma coluna numérica tem valor válido
    tem_dado_valido = False
    for coluna_numerica in estrutura['validacao_numerica']:
        if coluna_numerica in linha_dados:
            valor = linha_dados[coluna_numerica]
            try:
                if valor is not None and str(valor).strip():
                    float_val = float(valor)
                    if float_val >= 0:  # Valores numéricos válidos
                        tem_dado_valido = True
                        break
            except (ValueError, TypeError):
                continue
    
    # Para tabelas sem validação numérica, verificar se tem dados textuais válidos
    if not estrutura['validacao_numerica']:
        for coluna_norm, valor in linha_dados.items():
            if coluna_norm in ['linha_origem', 'tabela_tipo', 'aba_origem', 'valores_brutos']:
                continue
            if valor and str(valor).strip() and not validar_se_eh_header(valor, estrutura['colunas_esperadas']):
                tem_dado_valido = True
                break
    
    return tem_dado_valido

def extrair_tabela_corrigida(worksheet, estrutura, nome_tabela, aba_nome):
    """Extrai tabela com validação rigorosa contra headers"""
    
    linha_base, coluna_base = converter_coordenada_para_indices(estrutura['celula_base'])
    dados_validos = []
    
    # Extrair dados linha por linha, saltando a linha do header
    linha_atual = linha_base + 1  # Saltar o header
    max_linhas_vazias = 15
    linhas_vazias_consecutivas = 0
    
    while linhas_vazias_consecutivas < max_linhas_vazias and linha_atual <= worksheet.max_row:
        linha_dados = {}
        valores_brutos = []
        
        # Extrair valores de cada coluna
        for i, (coluna_esperada, coluna_norm) in enumerate(zip(estrutura['colunas_esperadas'], estrutura['colunas_normalizadas'])):
            coluna_atual = coluna_base + i
            letra_coluna = chr(ord('A') + coluna_atual - 1)
            celula_coord = f"{letra_coluna}{linha_atual}"
            
            try:
                valor = worksheet[celula_coord].value
            except:
                valor = None
            
            valores_brutos.append(valor)
            linha_dados[coluna_norm] = valor
        
        # Adicionar metadados
        linha_dados['linha_origem'] = linha_atual
        linha_dados['tabela_tipo'] = nome_tabela
        linha_dados['aba_origem'] = aba_nome
        linha_dados['valores_brutos'] = str(valores_brutos)
        
        # VALIDAÇÃO CRÍTICA: Verificar se é uma linha válida
        if validar_linha_dados(linha_dados, estrutura):
            dados_validos.append(linha_dados)
            linhas_vazias_consecutivas = 0
        else:
            linhas_vazias_consecutivas += 1
        
        linha_atual += 1
    
    return dados_validos

def extrair_metadados_aba(worksheet):
    """Extrai metadados da aba"""
    try:
        data_mov = worksheet['B1'].value
        loja_cel = worksheet['L1'].value
    except:
        data_mov = None
        loja_cel = None
    
    return {
        'data_movimento': data_mov,
        'loja_celula': loja_cel
    }

def processar_arquivo_corrigido(caminho_arquivo, loja_nome):
    """Processa arquivo com validação corrigida"""
    nome_arquivo = os.path.basename(caminho_arquivo)
    
    try:
        workbook = openpyxl.load_workbook(caminho_arquivo, read_only=True, data_only=True)
        
        dados_arquivo = {
            'arquivo': nome_arquivo,
            'loja': loja_nome,
            'tabelas_extraidas': {tipo: [] for tipo in ESTRUTURAS_TABELAS_CORRIGIDAS.keys()}
        }
        
        # Processar abas numéricas (dias)
        for sheet_name in workbook.sheetnames:
            if not re.match(r'^(0?[1-9]|[12]\d|3[01])$', str(sheet_name).strip()):
                continue
            
            worksheet = workbook[sheet_name]
            metadados_aba = extrair_metadados_aba(worksheet)
            
            # Extrair cada tipo de tabela
            for nome_tabela, estrutura in ESTRUTURAS_TABELAS_CORRIGIDAS.items():
                dados_tabela = extrair_tabela_corrigida(worksheet, estrutura, nome_tabela, sheet_name)
                
                # Adicionar metadados a cada linha válida
                for linha in dados_tabela:
                    linha.update(metadados_aba)
                    linha['loja_arquivo'] = loja_nome
                    linha['arquivo_origem'] = nome_arquivo
                
                dados_arquivo['tabelas_extraidas'][nome_tabela].extend(dados_tabela)
        
        workbook.close()
        return dados_arquivo
        
    except Exception as e:
        print(f"❌ Erro ao processar {nome_arquivo}: {str(e)}")
        return None

def reextrair_dados_caixa():
    """Re-extração completa dos dados do caixa com validação corrigida"""
    
    print("🔄 RE-EXTRAÇÃO CORRIGIDA DOS DADOS DO CAIXA")
    print("=" * 60)
    
    # Pastas
    pasta_cxs = r"d:\projetos\carne_facil\carne_facil\data\originais\cxs"
    pasta_extraidos = os.path.join(pasta_cxs, "extraidos_corrigidos")
    pasta_backup_antigo = os.path.join(pasta_cxs, "extraidos_por_tipo_backup")
    
    # Fazer backup da extração anterior
    pasta_anterior = os.path.join(pasta_cxs, "extraidos_por_tipo")
    if os.path.exists(pasta_anterior) and not os.path.exists(pasta_backup_antigo):
        os.rename(pasta_anterior, pasta_backup_antigo)
        print(f"📦 Backup da extração anterior: {pasta_backup_antigo}")
    
    # Criar nova estrutura
    os.makedirs(pasta_extraidos, exist_ok=True)
    for tipo in ESTRUTURAS_TABELAS_CORRIGIDAS.keys():
        os.makedirs(os.path.join(pasta_extraidos, tipo.lower()), exist_ok=True)
    
    # Processar cada loja
    lojas = ['maua', 'perus', 'rio_pequeno', 'sao_mateus', 'suzano', 'suzano2']
    dados_consolidados = {}
    
    for loja in lojas:
        print(f"\n🏪 PROCESSANDO: {loja.upper()}")
        print("-" * 40)
        
        pasta_loja = os.path.join(pasta_cxs, loja)
        if not os.path.exists(pasta_loja):
            print(f"❌ Pasta não encontrada: {pasta_loja}")
            continue
        
        arquivos = [f for f in os.listdir(pasta_loja) if f.endswith('.xlsx')]
        print(f"📄 {len(arquivos)} arquivos encontrados")
        
        dados_loja = {tipo: [] for tipo in ESTRUTURAS_TABELAS_CORRIGIDAS.keys()}
        
        for arquivo in sorted(arquivos):
            caminho_arquivo = os.path.join(pasta_loja, arquivo)
            dados_arquivo = processar_arquivo_corrigido(caminho_arquivo, loja)
            
            if dados_arquivo:
                for tipo, dados in dados_arquivo['tabelas_extraidas'].items():
                    dados_loja[tipo].extend(dados)
        
        # Salvar dados por loja e tipo
        total_registros = 0
        for tipo, dados in dados_loja.items():
            if dados:
                # CSV por loja
                df = pd.DataFrame(dados)
                arquivo_loja = os.path.join(pasta_extraidos, tipo.lower(), f"{tipo.lower()}_{loja}.csv")
                df.to_csv(arquivo_loja, index=False)
                
                total_registros += len(dados)
                print(f"  📊 {tipo}: {len(dados):,} registros → {arquivo_loja}")
        
        dados_consolidados[loja] = dados_loja
        print(f"✅ {loja.upper()}: {total_registros:,} registros totais")
    
    # Criar arquivos consolidados
    print(f"\n📋 CRIANDO ARQUIVOS CONSOLIDADOS")
    print("-" * 40)
    
    for tipo in ESTRUTURAS_TABELAS_CORRIGIDAS.keys():
        dados_tipo_consolidados = []
        
        for loja, dados_loja in dados_consolidados.items():
            dados_tipo_consolidados.extend(dados_loja[tipo])
        
        if dados_tipo_consolidados:
            df_consolidado = pd.DataFrame(dados_tipo_consolidados)
            arquivo_consolidado = os.path.join(pasta_extraidos, tipo.lower(), f"{tipo.lower()}_todas_lojas.csv")
            df_consolidado.to_csv(arquivo_consolidado, index=False)
            print(f"📊 {tipo}: {len(dados_tipo_consolidados):,} registros → {arquivo_consolidado}")
    
    # Renomear pasta final
    pasta_final = os.path.join(pasta_cxs, "extraidos_por_tipo")
    if os.path.exists(pasta_extraidos):
        os.rename(pasta_extraidos, pasta_final)
    
    print(f"\n🎉 RE-EXTRAÇÃO CONCLUÍDA!")
    print(f"📁 Dados corrigidos em: {pasta_final}")
    print(f"📦 Backup anterior em: {pasta_backup_antigo}")

if __name__ == "__main__":
    reextrair_dados_caixa()