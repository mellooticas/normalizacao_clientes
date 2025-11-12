"""
Verifica se os arquivos Excel têm Named Ranges (Gerenciador de Nomes)
"""

import openpyxl
from pathlib import Path

# Arquivo de exemplo
arquivo = Path('dados_processados/originais/cxs/planilhas_originais/maua/jan_24.xlsx')

print('=' * 80)
print('ANÁLISE DE GERENCIADOR DE NOMES (NAMED RANGES)')
print('=' * 80)
print(f'\nArquivo: {arquivo.name}')

# Carregar workbook
wb = openpyxl.load_workbook(arquivo, data_only=False)

print(f'\n📋 Abas: {wb.sheetnames}')

# Verificar Named Ranges
print(f'\n{"=" * 80}')
print('GERENCIADOR DE NOMES (NAMED RANGES)')
print('=' * 80)

if wb.defined_names:
    print(f'\n✅ Total de nomes definidos: {len(wb.defined_names.definedName)}')
    
    for idx, (name, defn) in enumerate(wb.defined_names.items(), 1):
        print(f'\n{idx}. Nome: {name}')
        print(f'   Definição: {defn.value}')
        
        # Tentar obter mais detalhes
        if hasattr(defn, 'destinations'):
            print(f'   Destinos: {list(defn.destinations)}')
        
        if idx >= 10:  # Mostrar apenas os primeiros 10
            restantes = len(wb.defined_names.definedName) - 10
            if restantes > 0:
                print(f'\n... e mais {restantes} nomes definidos')
            break
else:
    print('\n❌ Nenhum nome definido encontrado')

# Verificar uma aba específica
print(f'\n{"=" * 80}')
print('ANÁLISE DA ABA "01" (DIA 1)')
print('=' * 80)

if '01' in wb.sheetnames:
    ws = wb['01']
    
    print(f'\nDimensões: {ws.dimensions}')
    print(f'Máxima linha: {ws.max_row}')
    print(f'Máxima coluna: {ws.max_column}')
    
    # Ler primeiras linhas para ver estrutura
    print(f'\n📄 Primeiras 10 linhas:')
    for row_idx in range(1, 11):
        row_values = []
        for col_idx in range(1, min(10, ws.max_column + 1)):
            cell = ws.cell(row_idx, col_idx)
            if cell.value is not None:
                row_values.append(f'{cell.value}')
        
        if row_values:
            print(f'   Linha {row_idx}: {" | ".join(row_values[:5])}')

wb.close()

print(f'\n{"=" * 80}')
print('✅ Análise concluída!')
print('=' * 80)
