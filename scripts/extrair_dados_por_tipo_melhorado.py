#!/usr/bin/env python3
"""
Sistema avançado de extração e organização das tabelas de caixa por tipo
Sistema Carne Fácil - Análise detalhada e separação por tipo de tabela
"""

import pandas as pd
import openpyxl
import os
import json
from datetime import datetime
import re

# Definição das estruturas de tabelas detalhadas
ESTRUTURAS_TABELAS = {
    'VENDAS': {
        'celula_base': 'E5',
        'colunas': ['Nº Venda', 'Cliente', 'Forma de Pgto', 'Valor Venda', 'Entrada'],
        'tipo': 'financeiro',
        'descricao': 'Vendas realizadas no dia'
    },
    'RESTANTE_ENTRADA': {
        'celula_base': 'E25', 
        'colunas': ['Nº Venda', 'Cliente', 'Forma de Pgto', 'Valor Venda', 'Entrada'],
        'tipo': 'financeiro',
        'descricao': 'Valores restantes de entrada'
    },
    'RECEBIMENTO_CARNE': {
        'celula_base': 'E34',
        'colunas': ['OS', 'Cliente', 'Forma de Pgto', 'Valor Parcela', 'Nº Parcela'],
        'tipo': 'recebimento',
        'descricao': 'Recebimentos de parcelas de carnê'
    },
    'OS_ENTREGUES_DIA': {
        'celula_base': 'K14',
        'colunas': ['OS', 'Vendedor', 'CARNE'],
        'tipo': 'operacional',
        'descricao': 'OS entregues no dia'
    },
    'ENTREGA_CARNE': {
        'celula_base': 'K34',
        'colunas': ['OS', 'Parcelas', 'Valor Total'],
        'tipo': 'entrega',
        'descricao': 'Carnês entregues'
    }
}

def converter_coordenada_para_indices(coordenada):
    """
    Converte coordenada Excel (ex: E5) para índices (linha, coluna) 
    """
    from openpyxl.utils import coordinate_to_tuple
    return coordinate_to_tuple(coordenada)

def extrair_dados_celula(worksheet, coordenada):
    """
    Extrai dados de uma célula específica
    """
    try:
        valor = worksheet[coordenada].value
        return valor if valor is not None else ""
    except:
        return ""

def normalizar_nome_coluna(nome):
    """
    Normaliza nomes de colunas para uso em DataFrames
    """
    return nome.lower().replace(' ', '_').replace('º', 'n').replace('ª', 'a')

def extrair_tabela_dinamica(worksheet, estrutura, nome_tabela, aba_nome):
    """
    Extrai uma tabela dinâmica baseada na estrutura definida
    """
    linha_base, coluna_base = converter_coordenada_para_indices(estrutura['celula_base'])
    dados_tabela = []
    
    # Verificar se existe dados na posição base
    primeira_celula = extrair_dados_celula(worksheet, estrutura['celula_base'])
    
    if not primeira_celula and not str(primeira_celula).strip():
        return []
    
    # Definir colunas esperadas
    colunas_esperadas = estrutura['colunas']
    
    # Extrair dados linha por linha a partir da linha seguinte ao cabeçalho
    linha_atual = linha_base + 1
    max_linhas_vazias = 10  # Parar após 10 linhas vazias consecutivas
    linhas_vazias_consecutivas = 0
    
    while linhas_vazias_consecutivas < max_linhas_vazias and linha_atual <= min(worksheet.max_row, linha_base + 100):
        linha_dados = {}
        linha_tem_dados = False
        valores_linha = []
        
        # Extrair dados de cada coluna
        for i, coluna_nome in enumerate(colunas_esperadas):
            coluna_atual = coluna_base + i
            
            # Calcular coordenada da célula
            letra_coluna = chr(ord('A') + coluna_atual - 1)
            celula_coord = f"{letra_coluna}{linha_atual}"
            
            valor = extrair_dados_celula(worksheet, celula_coord)
            valores_linha.append(valor)
            
            # Normalizar nome da coluna
            coluna_normalizada = normalizar_nome_coluna(coluna_nome)
            linha_dados[coluna_normalizada] = valor
            
            if valor and str(valor).strip():
                linha_tem_dados = True
        
        if linha_tem_dados:
            # Adicionar metadados
            linha_dados['linha_origem'] = linha_atual
            linha_dados['tabela_tipo'] = nome_tabela
            linha_dados['aba_origem'] = aba_nome
            linha_dados['valores_brutos'] = str(valores_linha)
            
            dados_tabela.append(linha_dados)
            linhas_vazias_consecutivas = 0
        else:
            linhas_vazias_consecutivas += 1
        
        linha_atual += 1
    
    return dados_tabela

def extrair_metadados_aba(worksheet):
    """
    Extrai metadados da aba (data e loja)
    """
    metadados = {
        'data_movimento': extrair_dados_celula(worksheet, 'B1'),
        'loja_celula': extrair_dados_celula(worksheet, 'L1')
    }
    
    return metadados

def processar_arquivo_caixa(caminho_arquivo, loja_nome):
    """
    Processa um arquivo de caixa completo extraindo todas as tabelas
    """
    nome_arquivo = os.path.basename(caminho_arquivo)
    print(f"      📄 {nome_arquivo}")
    
    try:
        workbook = openpyxl.load_workbook(caminho_arquivo, read_only=True, data_only=True)
        dados_arquivo = {
            'arquivo': nome_arquivo,
            'loja': loja_nome,
            'abas_processadas': [],
            'tabelas_extraidas': {tipo: [] for tipo in ESTRUTURAS_TABELAS.keys()},
            'metadados': []
        }
        
        # Processar cada aba (dia do mês)
        for sheet_name in workbook.sheetnames:
            # Apenas abas numéricas (dias do mês 1-31)
            if not re.match(r'^(0?[1-9]|[12]\d|3[01])$', str(sheet_name).strip()):
                continue
                
            worksheet = workbook[sheet_name]
            dados_arquivo['abas_processadas'].append(sheet_name)
            
            # Extrair metadados da aba
            metadados_aba = extrair_metadados_aba(worksheet)
            
            # Processar cada tipo de tabela
            for nome_tabela, estrutura in ESTRUTURAS_TABELAS.items():
                dados_tabela = extrair_tabela_dinamica(worksheet, estrutura, nome_tabela, sheet_name)
                
                # Adicionar metadados a cada linha
                for linha in dados_tabela:
                    linha.update(metadados_aba)
                    linha['loja_arquivo'] = loja_nome
                    linha['arquivo_origem'] = nome_arquivo
                
                dados_arquivo['tabelas_extraidas'][nome_tabela].extend(dados_tabela)
            
            # Salvar metadados da aba
            metadados_aba.update({
                'aba': sheet_name,
                'arquivo': nome_arquivo,
                'loja': loja_nome
            })
            dados_arquivo['metadados'].append(metadados_aba)
        
        workbook.close()
        
        # Relatório do arquivo
        total_registros = sum(len(tabela) for tabela in dados_arquivo['tabelas_extraidas'].values())
        if total_registros > 0:
            print(f"         ✅ {len(dados_arquivo['abas_processadas'])} abas, {total_registros} registros")
        else:
            print(f"         ⚠️  {len(dados_arquivo['abas_processadas'])} abas, nenhum registro")
        
        return dados_arquivo
        
    except Exception as e:
        print(f"         ❌ Erro: {str(e)}")
        return None

def processar_loja(loja_nome):
    """
    Processa todos os arquivos de uma loja
    """
    print(f"\n🏪 LOJA: {loja_nome.upper()}")
    print("-" * 30)
    
    pasta_loja = f'data/originais/cxs/{loja_nome}'
    
    if not os.path.exists(pasta_loja):
        print(f"   ❌ Pasta não encontrada: {pasta_loja}")
        return None
    
    # Listar arquivos Excel da loja
    arquivos = [f for f in os.listdir(pasta_loja) if f.endswith('.xlsx')]
    
    if not arquivos:
        print(f"   ⚠️  Nenhum arquivo Excel encontrado")
        return None
    
    print(f"   📋 {len(arquivos)} arquivos encontrados")
    
    # Dados consolidados da loja
    dados_loja = {
        'loja': loja_nome,
        'total_arquivos': len(arquivos),
        'arquivos_processados': [],
        'tabelas_consolidadas': {tipo: [] for tipo in ESTRUTURAS_TABELAS.keys()},
        'estatisticas': {},
        'metadados_arquivos': []
    }
    
    # Processar cada arquivo
    for arquivo in sorted(arquivos):
        caminho_arquivo = os.path.join(pasta_loja, arquivo)
        dados_arquivo = processar_arquivo_caixa(caminho_arquivo, loja_nome)
        
        if dados_arquivo:
            dados_loja['arquivos_processados'].append(arquivo)
            dados_loja['metadados_arquivos'].append(dados_arquivo)
            
            # Consolidar tabelas
            for tipo_tabela, dados_tabela in dados_arquivo['tabelas_extraidas'].items():
                dados_loja['tabelas_consolidadas'][tipo_tabela].extend(dados_tabela)
    
    # Calcular estatísticas
    for tipo_tabela, dados_tabela in dados_loja['tabelas_consolidadas'].items():
        dados_loja['estatisticas'][tipo_tabela] = len(dados_tabela)
    
    # Relatório da loja
    print(f"   ✅ {len(dados_loja['arquivos_processados'])}/{len(arquivos)} arquivos processados")
    for tipo, quantidade in dados_loja['estatisticas'].items():
        if quantidade > 0:
            print(f"      📊 {tipo}: {quantidade} registros")
    
    return dados_loja

def criar_estrutura_pastas():
    """
    Cria estrutura de pastas para organizar dados por tipo de tabela
    """
    print(f"\n📁 CRIANDO ESTRUTURA DE PASTAS")
    print("=" * 40)
    
    # Pasta base para dados extraídos
    pasta_base = 'data/originais/cxs/extraidos_por_tipo'
    os.makedirs(pasta_base, exist_ok=True)
    
    # Pastas por tipo de tabela
    for tipo_tabela, estrutura in ESTRUTURAS_TABELAS.items():
        pasta_tipo = os.path.join(pasta_base, tipo_tabela.lower())
        os.makedirs(pasta_tipo, exist_ok=True)
        print(f"   📂 {tipo_tabela}: {estrutura['descricao']}")
    
    # Pasta para análises consolidadas
    pasta_analises = os.path.join(pasta_base, '_analises')
    os.makedirs(pasta_analises, exist_ok=True)
    print(f"   📊 Análises: {pasta_analises}")
    
    return pasta_base

def salvar_dados_por_tipo(dados_todas_lojas, pasta_base):
    """
    Salva os dados organizados por tipo de tabela
    """
    print(f"\n💾 SALVANDO DADOS POR TIPO DE TABELA")
    print("=" * 50)
    
    for tipo_tabela, estrutura in ESTRUTURAS_TABELAS.items():
        print(f"\n   📋 {tipo_tabela} - {estrutura['descricao']}")
        
        # Consolidar dados de todas as lojas para este tipo
        dados_consolidados = []
        
        for loja_nome, dados_loja in dados_todas_lojas.items():
            if dados_loja and tipo_tabela in dados_loja['tabelas_consolidadas']:
                dados_tipo = dados_loja['tabelas_consolidadas'][tipo_tabela]
                dados_consolidados.extend(dados_tipo)
                
                if dados_tipo:
                    print(f"      🏪 {loja_nome}: {len(dados_tipo)} registros")
        
        if dados_consolidados:
            # Pasta do tipo
            pasta_tipo = os.path.join(pasta_base, tipo_tabela.lower())
            
            # DataFrame consolidado
            df_consolidado = pd.DataFrame(dados_consolidados)
            
            # CSV consolidado (todas as lojas)
            caminho_csv_consolidado = os.path.join(pasta_tipo, f'{tipo_tabela.lower()}_todas_lojas.csv')
            df_consolidado.to_csv(caminho_csv_consolidado, index=False, encoding='utf-8')
            
            # CSVs individuais por loja
            for loja_nome in dados_todas_lojas.keys():
                dados_loja_tipo = [d for d in dados_consolidados if d.get('loja_arquivo') == loja_nome]
                if dados_loja_tipo:
                    df_loja = pd.DataFrame(dados_loja_tipo)
                    caminho_csv_loja = os.path.join(pasta_tipo, f'{tipo_tabela.lower()}_{loja_nome}.csv')
                    df_loja.to_csv(caminho_csv_loja, index=False, encoding='utf-8')
            
            # JSON com estrutura detalhada
            dados_json = {
                'tipo_tabela': tipo_tabela,
                'estrutura': estrutura,
                'total_registros': len(dados_consolidados),
                'lojas_com_dados': list(set(d.get('loja_arquivo') for d in dados_consolidados)),
                'colunas_encontradas': list(df_consolidado.columns),
                'amostra_dados': dados_consolidados[:5] if dados_consolidados else []
            }
            
            caminho_json = os.path.join(pasta_tipo, f'{tipo_tabela.lower()}_estrutura.json')
            with open(caminho_json, 'w', encoding='utf-8') as f:
                json.dump(dados_json, f, indent=2, ensure_ascii=False, default=str)
            
            print(f"      💾 Total: {len(dados_consolidados)} registros salvos")
        else:
            print(f"      ⚠️  Nenhum dado encontrado")

def gerar_relatorio_consolidado(dados_todas_lojas, pasta_base):
    """
    Gera relatório consolidado de todas as extrações
    """
    print(f"\n📊 GERANDO RELATÓRIO CONSOLIDADO")
    print("=" * 40)
    
    relatorio = {
        'data_processamento': datetime.now().isoformat(),
        'versao_sistema': '2.0',
        'estruturas_tabelas_definidas': ESTRUTURAS_TABELAS,
        'resumo_geral': {
            'total_lojas_processadas': len([l for l in dados_todas_lojas.values() if l]),
            'total_arquivos': sum(len(l['arquivos_processados']) for l in dados_todas_lojas.values() if l),
            'total_registros': 0
        },
        'resumo_por_loja': {},
        'resumo_por_tipo': {},
        'detalhes_por_tipo': {}
    }
    
    # Resumo por loja
    for loja_nome, dados_loja in dados_todas_lojas.items():
        if dados_loja:
            total_registros_loja = sum(dados_loja['estatisticas'].values())
            relatorio['resumo_por_loja'][loja_nome] = {
                'arquivos_processados': len(dados_loja['arquivos_processados']),
                'total_registros': total_registros_loja,
                'estatisticas_por_tipo': dados_loja['estatisticas']
            }
            relatorio['resumo_geral']['total_registros'] += total_registros_loja
    
    # Resumo por tipo
    for tipo_tabela, estrutura in ESTRUTURAS_TABELAS.items():
        total_tipo = sum(
            dados_loja['estatisticas'].get(tipo_tabela, 0) 
            for dados_loja in dados_todas_lojas.values() 
            if dados_loja
        )
        
        lojas_com_dados = [
            loja for loja, dados_loja in dados_todas_lojas.items()
            if dados_loja and dados_loja['estatisticas'].get(tipo_tabela, 0) > 0
        ]
        
        relatorio['resumo_por_tipo'][tipo_tabela] = total_tipo
        relatorio['detalhes_por_tipo'][tipo_tabela] = {
            'total_registros': total_tipo,
            'lojas_com_dados': lojas_com_dados,
            'estrutura_definida': estrutura,
            'distribuicao_por_loja': {
                loja: dados_loja['estatisticas'].get(tipo_tabela, 0)
                for loja, dados_loja in dados_todas_lojas.items()
                if dados_loja and dados_loja['estatisticas'].get(tipo_tabela, 0) > 0
            }
        }
    
    # Salvar relatório
    pasta_analises = os.path.join(pasta_base, '_analises')
    caminho_relatorio = os.path.join(pasta_analises, 'relatorio_extracao_por_tipos.json')
    
    with open(caminho_relatorio, 'w', encoding='utf-8') as f:
        json.dump(relatorio, f, indent=2, ensure_ascii=False, default=str)
    
    # Exibir resumo
    print(f"   📈 Lojas processadas: {relatorio['resumo_geral']['total_lojas_processadas']}")
    print(f"   📄 Arquivos processados: {relatorio['resumo_geral']['total_arquivos']}")
    print(f"   📋 Total de registros: {relatorio['resumo_geral']['total_registros']}")
    
    print(f"\n   📊 Registros por tipo:")
    for tipo, total in relatorio['resumo_por_tipo'].items():
        if total > 0:
            lojas = relatorio['detalhes_por_tipo'][tipo]['lojas_com_dados']
            print(f"      📋 {tipo}: {total} registros ({len(lojas)} lojas)")
    
    print(f"\n💾 Relatório completo: {caminho_relatorio}")
    
    return relatorio

def main():
    """Função principal"""
    print("🔄 SISTEMA AVANÇADO DE EXTRAÇÃO DE TABELAS DE CAIXA")
    print("=" * 80)
    print("📋 Tipos de tabelas a extrair:")
    for tipo, estrutura in ESTRUTURAS_TABELAS.items():
        print(f"   • {tipo}: {estrutura['descricao']} (célula {estrutura['celula_base']})")
    
    # Criar estrutura de pastas
    pasta_base = criar_estrutura_pastas()
    
    # Processar todas as lojas
    lojas = ['maua', 'perus', 'rio_pequeno', 'sao_mateus', 'suzano', 'suzano2']
    dados_todas_lojas = {}
    
    print(f"\n🏪 PROCESSANDO {len(lojas)} LOJAS")
    print("=" * 40)
    
    for loja in lojas:
        dados_loja = processar_loja(loja)
        dados_todas_lojas[loja] = dados_loja
    
    # Salvar dados por tipo
    salvar_dados_por_tipo(dados_todas_lojas, pasta_base)
    
    # Gerar relatório consolidado
    relatorio = gerar_relatorio_consolidado(dados_todas_lojas, pasta_base)
    
    print(f"\n✅ EXTRAÇÃO POR TIPOS CONCLUÍDA!")
    print(f"📁 Dados organizados em: {pasta_base}")
    print(f"📊 {relatorio['resumo_geral']['total_registros']} registros extraídos")
    print(f"🎯 Próximo passo: Análise dos dados por tipo de tabela")

if __name__ == "__main__":
    main()