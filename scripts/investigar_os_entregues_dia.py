#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Re-extração específica para OS_ENTREGUES_DIA
Corrigindo problema de extração
"""

import pandas as pd
import openpyxl
import os
from datetime import datetime

def investigar_os_entregues_dia():
    """Investiga e re-extrai dados de OS_ENTREGUES_DIA"""
    
    print("🔍 INVESTIGAÇÃO E RE-EXTRAÇÃO: OS_ENTREGUES_DIA")
    print("=" * 60)
    
    # Configuração da tabela OS_ENTREGUES_DIA
    config_os_entregues = {
        'celula_base': 'K14',
        'colunas_esperadas': ['OS', 'Vendedor', 'CARNE'],
        'colunas_normalizadas': ['os', 'vendedor', 'carne']
    }
    
    # Testar em um arquivo específico primeiro
    pasta_maua = r"d:\projetos\carne_facil\carne_facil\data\originais\cxs\maua"
    arquivos_maua = [f for f in os.listdir(pasta_maua) if f.endswith('.xlsx')]
    
    if not arquivos_maua:
        print("❌ Nenhum arquivo Excel encontrado em MAUA")
        return
    
    arquivo_teste = os.path.join(pasta_maua, arquivos_maua[0])
    print(f"🧪 TESTANDO ARQUIVO: {arquivos_maua[0]}")
    print("-" * 40)
    
    try:
        workbook = openpyxl.load_workbook(arquivo_teste, read_only=True, data_only=True)
        
        # Testar algumas abas
        abas_teste = [sheet for sheet in workbook.sheetnames if sheet.isdigit()][:3]
        
        for aba in abas_teste:
            print(f"\n📋 ABA: {aba}")
            worksheet = workbook[aba]
            
            # Verificar célula base K14
            celula_k14 = worksheet['K14'].value
            print(f"   K14: {celula_k14}")
            
            # Verificar área ao redor de K14
            print("   Área K14-M20:")
            for linha in range(14, 21):
                for coluna in ['K', 'L', 'M']:
                    try:
                        valor = worksheet[f'{coluna}{linha}'].value
                        if valor:
                            print(f"     {coluna}{linha}: {valor}")
                    except:
                        pass
            
            # Verificar se há padrão de dados
            dados_encontrados = []
            linha_atual = 15  # Começar após o possível header
            
            while linha_atual <= 30:  # Testar até linha 30
                linha_dados = []
                for i, coluna in enumerate(['K', 'L', 'M']):
                    try:
                        valor = worksheet[f'{coluna}{linha_atual}'].value
                        linha_dados.append(valor)
                    except:
                        linha_dados.append(None)
                
                # Se encontrou dados válidos
                if any(v for v in linha_dados if v and str(v).strip()):
                    dados_encontrados.append(f"   Linha {linha_atual}: {linha_dados}")
                    if len(dados_encontrados) >= 5:  # Mostrar apenas primeiros 5
                        break
                
                linha_atual += 1
            
            if dados_encontrados:
                print("   📊 DADOS ENCONTRADOS:")
                for dado in dados_encontrados:
                    print(dado)
            else:
                print("   ⚠️  Nenhum dado encontrado")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Erro ao processar arquivo: {str(e)}")
    
    # Testar diferentes posições possíveis
    print(f"\n🔍 TESTANDO POSIÇÕES ALTERNATIVAS")
    print("-" * 40)
    
    posicoes_teste = [
        {'celula': 'K14', 'desc': 'Posição original'},
        {'celula': 'K15', 'desc': 'Uma linha abaixo'},
        {'celula': 'J14', 'desc': 'Uma coluna à esquerda'},
        {'celula': 'L14', 'desc': 'Uma coluna à direita'},
        {'celula': 'K13', 'desc': 'Uma linha acima'}
    ]
    
    try:
        workbook = openpyxl.load_workbook(arquivo_teste, read_only=True, data_only=True)
        worksheet = workbook[abas_teste[0]]  # Usar primeira aba
        
        for pos in posicoes_teste:
            try:
                valor = worksheet[pos['celula']].value
                print(f"   {pos['celula']} ({pos['desc']}): {valor}")
            except:
                print(f"   {pos['celula']} ({pos['desc']}): ERRO")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Erro na verificação de posições: {str(e)}")

def reextrair_os_entregues_corrigido():
    """Re-extrai OS_ENTREGUES_DIA com configuração corrigida"""
    
    print(f"\n🔄 RE-EXTRAÇÃO CORRIGIDA: OS_ENTREGUES_DIA")
    print("=" * 60)
    
    # Configuração corrigida (pode ser ajustada baseado na investigação)
    config_corrigida = {
        'celula_base': 'K15',  # Testando uma linha abaixo
        'colunas_esperadas': ['OS', 'Vendedor', 'CARNE'],
        'colunas_normalizadas': ['os', 'vendedor', 'carne']
    }
    
    pasta_base_cxs = r"d:\projetos\carne_facil\carne_facil\data\originais\cxs"
    pasta_saida = os.path.join(pasta_base_cxs, "extraidos_corrigidos", "os_entregues_dia")
    
    lojas = ['maua', 'perus', 'rio_pequeno', 'sao_mateus', 'suzano', 'suzano2']
    
    dados_consolidados = []
    
    for loja in lojas:
        print(f"\n🏪 PROCESSANDO: {loja.upper()}")
        print("-" * 30)
        
        pasta_loja = os.path.join(pasta_base_cxs, loja)
        arquivos = [f for f in os.listdir(pasta_loja) if f.endswith('.xlsx')]
        
        dados_loja = []
        
        for arquivo in sorted(arquivos):
            caminho_arquivo = os.path.join(pasta_loja, arquivo)
            
            try:
                workbook = openpyxl.load_workbook(caminho_arquivo, read_only=True, data_only=True)
                
                for sheet_name in workbook.sheetnames:
                    if not sheet_name.isdigit():
                        continue
                    
                    worksheet = workbook[sheet_name]
                    
                    # Extrair metadados da aba
                    try:
                        data_mov = worksheet['B1'].value
                        loja_cel = worksheet['L1'].value
                    except:
                        data_mov = None
                        loja_cel = None
                    
                    # Extrair dados da tabela OS_ENTREGUES_DIA
                    linha_base = 15  # Começar da linha 15 (K15)
                    
                    for linha_atual in range(linha_base, linha_base + 50):  # Testar 50 linhas
                        linha_dados = {}
                        valores_brutos = []
                        tem_dados = False
                        
                        # Extrair valores das 3 colunas K, L, M
                        for i, (col_esp, col_norm) in enumerate(zip(config_corrigida['colunas_esperadas'], config_corrigida['colunas_normalizadas'])):
                            coluna_letra = chr(ord('K') + i)  # K, L, M
                            celula_coord = f"{coluna_letra}{linha_atual}"
                            
                            try:
                                valor = worksheet[celula_coord].value
                            except:
                                valor = None
                            
                            valores_brutos.append(valor)
                            linha_dados[col_norm] = valor
                            
                            if valor and str(valor).strip() and not str(valor).strip() in ['OS', 'Vendedor', 'CARNE']:
                                tem_dados = True
                        
                        if tem_dados:
                            # Adicionar metadados
                            linha_dados.update({
                                'linha_origem': linha_atual,
                                'tabela_tipo': 'OS_ENTREGUES_DIA',
                                'aba_origem': sheet_name,
                                'valores_brutos': str(valores_brutos),
                                'data_movimento': data_mov,
                                'loja_celula': loja_cel,
                                'loja_arquivo': loja,
                                'arquivo_origem': arquivo
                            })
                            
                            dados_loja.append(linha_dados)
                
                workbook.close()
                
            except Exception as e:
                print(f"   ❌ Erro em {arquivo}: {str(e)}")
        
        # Salvar dados da loja
        if dados_loja:
            df_loja = pd.DataFrame(dados_loja)
            arquivo_loja = os.path.join(pasta_saida, f"os_entregues_dia_{loja}.csv")
            df_loja.to_csv(arquivo_loja, index=False)
            print(f"   ✅ {len(dados_loja)} registros → {arquivo_loja}")
            dados_consolidados.extend(dados_loja)
        else:
            print(f"   ⚠️  Nenhum dado encontrado")
    
    # Salvar consolidado
    if dados_consolidados:
        df_consolidado = pd.DataFrame(dados_consolidados)
        arquivo_consolidado = os.path.join(pasta_saida, "os_entregues_dia_todas_lojas.csv")
        df_consolidado.to_csv(arquivo_consolidado, index=False)
        print(f"\n📊 CONSOLIDADO: {len(dados_consolidados)} registros → {arquivo_consolidado}")
    
    print(f"\n🎉 RE-EXTRAÇÃO CONCLUÍDA!")
    print(f"📊 Total de registros: {len(dados_consolidados)}")

if __name__ == "__main__":
    investigar_os_entregues_dia()
    
    resposta = input("\nDeseja continuar com a re-extração? (s/n): ")
    if resposta.lower() == 's':
        reextrair_os_entregues_corrigido()