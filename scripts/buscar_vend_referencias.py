"""
Busca por Named Ranges e células que contenham "vend"
"""

import openpyxl
from pathlib import Path

# Arquivo de exemplo
arquivo = Path('dados_processados/originais/cxs/planilhas_originais/maua/jan_24.xlsx')

print('=' * 80)
print('BUSCA POR "VEND" - NAMED RANGES E CÉLULAS')
print('=' * 80)
print(f'\nArquivo: {arquivo.name}')

# Carregar workbook
wb = openpyxl.load_workbook(arquivo, data_only=False)

print(f'\n📋 Total de abas: {len(wb.sheetnames)}')

# 1. Buscar Named Ranges com "vend"
print(f'\n{"=" * 80}')
print('1. NAMED RANGES CONTENDO "VEND"')
print('=' * 80)

if wb.defined_names:
    nomes_vend = []
    for name, defn in wb.defined_names.items():
        if 'vend' in name.lower():
            nomes_vend.append((name, defn.value))
    
    if nomes_vend:
        print(f'\n✅ Encontrados {len(nomes_vend)} nomes com "vend":')
        for name, value in nomes_vend:
            print(f'\n   Nome: {name}')
            print(f'   Valor: {value}')
    else:
        print('\n❌ Nenhum Named Range com "vend" encontrado')
else:
    print('\n❌ Arquivo não possui Named Ranges')

# 2. Buscar células contendo "vend" em todas as abas
print(f'\n{"=" * 80}')
print('2. CÉLULAS CONTENDO "VEND" (PRIMEIRAS 3 ABAS)')
print('=' * 80)

abas_para_analisar = ['resumo_cx', '01', '02']

for aba_nome in abas_para_analisar:
    if aba_nome not in wb.sheetnames:
        continue
    
    print(f'\n--- ABA: {aba_nome} ---')
    ws = wb[aba_nome]
    
    celulas_encontradas = []
    
    # Buscar em todas as células (limitado a 100 linhas)
    for row_idx in range(1, min(101, ws.max_row + 1)):
        for col_idx in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row_idx, col_idx)
            
            if cell.value is not None:
                cell_str = str(cell.value).lower()
                if 'vend' in cell_str:
                    celulas_encontradas.append({
                        'linha': row_idx,
                        'coluna': col_idx,
                        'letra': openpyxl.utils.get_column_letter(col_idx),
                        'valor': cell.value
                    })
    
    if celulas_encontradas:
        print(f'✅ Encontradas {len(celulas_encontradas)} células com "vend":')
        for cel in celulas_encontradas[:10]:
            print(f'   {cel["letra"]}{cel["linha"]}: {cel["valor"]}')
        
        if len(celulas_encontradas) > 10:
            print(f'   ... e mais {len(celulas_encontradas) - 10} células')
    else:
        print('❌ Nenhuma célula com "vend" encontrada')

# 3. Analisar aba '01' especificamente - linha 5 (cabeçalho)
print(f'\n{"=" * 80}')
print('3. ANÁLISE DETALHADA DA ABA "01" - LINHA 5 (CABEÇALHO)')
print('=' * 80)

if '01' in wb.sheetnames:
    ws = wb['01']
    
    print('\n📋 Linha 5 completa (possível cabeçalho):')
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(5, col_idx)
        if cell.value is not None:
            letra = openpyxl.utils.get_column_letter(col_idx)
            print(f'   {letra}5: {cell.value}')

# 4. Buscar na aba 'base'
print(f'\n{"=" * 80}')
print('4. ANÁLISE DA ABA "BASE"')
print('=' * 80)

if 'base' in wb.sheetnames:
    ws = wb['base']
    
    print(f'\n📊 Dimensões: {ws.dimensions}')
    print(f'\n📋 Primeiras 10 linhas:')
    
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_values = []
        for col_idx in range(1, min(10, ws.max_column + 1)):
            cell = ws.cell(row_idx, col_idx)
            if cell.value is not None:
                row_values.append(str(cell.value))
        
        if row_values:
            print(f'   Linha {row_idx}: {" | ".join(row_values)}')

wb.close()

print(f'\n{"=" * 80}')
print('✅ Busca concluída!')
print('=' * 80)
