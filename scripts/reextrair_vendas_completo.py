"""
Reextracao completa de vendas incluindo multiplas formas de pagamento
Versao sem emojis para compatibilidade com terminal Windows
"""

import openpyxl
import pandas as pd
from pathlib import Path
import re

base_dir = Path('dados_processados/originais/cxs/planilhas_originais')
lojas = ['maua', 'perus', 'rio_pequeno', 'sao_mateus', 'suzano', 'suzano2']

print('=' * 80)
print('REEXTRACAO COMPLETA - COM MULTIPLAS FORMAS DE PAGAMENTO')
print('=' * 80)

def extrair_mes_ano_arquivo(nome_arquivo):
    meses = {
        'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
        'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
    }
    match = re.match(r'([a-z]{3})_(\d{2})\.xlsx', nome_arquivo.lower())
    if match:
        mes_str, ano_str = match.groups()
        mes = meses.get(mes_str, 1)
        ano = 2000 + int(ano_str)
        return mes, ano
    return None, None

def encontrar_linha_limite(ws):
    for row_idx in range(1, min(100, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value and 'restante entrada' in str(cell.value).lower():
                return row_idx
    return 50

for loja in lojas:
    loja_dir = base_dir / loja
    if not loja_dir.exists():
        continue
    
    print(f'\n{"=" * 80}')
    print(f'LOJA: {loja.upper()}')
    print('=' * 80)
    
    arquivos = sorted(loja_dir.glob('*.xlsx'))
    print(f'\nTotal de arquivos: {len(arquivos)}')
    
    todas_vendas = []
    total_principais = 0
    total_adicionais = 0
    
    for arquivo in arquivos:
        print(f'\nProcessando: {arquivo.name}')
        
        mes, ano = extrair_mes_ano_arquivo(arquivo.name)
        if not mes or not ano:
            print(f'   ERRO: Nao conseguiu extrair mes/ano')
            continue
        
        print(f'   Periodo: {mes:02d}/{ano}')
        
        try:
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            vendas_arquivo = 0
            adicionais_arquivo = 0
            
            for dia in range(1, 32):
                aba_nome = f'{dia:02d}'
                if aba_nome not in wb.sheetnames:
                    continue
                
                ws = wb[aba_nome]
                linha_limite = encontrar_linha_limite(ws)
                
                ultima_venda = None
                ultima_cliente = None
                data_venda = f'{ano}-{mes:02d}-{dia:02d}'
                
                for row_idx in range(6, linha_limite):
                    num_venda = ws.cell(row_idx, 5).value
                    cliente = ws.cell(row_idx, 6).value
                    forma_pgto = ws.cell(row_idx, 7).value
                    valor_venda = ws.cell(row_idx, 8).value
                    entrada = ws.cell(row_idx, 9).value
                    
                    # Linha PRINCIPAL (tem OS e Cliente)
                    if num_venda and cliente:
                        if str(num_venda).strip().lower() not in ['nº venda', 'vendas', 'restante entrada']:
                            try:
                                ultima_venda = str(num_venda).strip()
                                ultima_cliente = str(cliente).strip()
                                
                                todas_vendas.append({
                                    'data_movimento': data_venda,
                                    'nn_venda': ultima_venda,
                                    'cliente': ultima_cliente,
                                    'forma_de_pgto': str(forma_pgto).strip() if forma_pgto else '',
                                    'valor_venda': valor_venda if valor_venda else 0,
                                    'entrada': entrada if entrada else 0
                                })
                                vendas_arquivo += 1
                            except:
                                pass
                    
                    # Linha ADICIONAL (sem OS/Cliente mas com pagamento)
                    elif (not num_venda or not cliente) and (forma_pgto or entrada):
                        if ultima_venda and ultima_cliente:
                            try:
                                todas_vendas.append({
                                    'data_movimento': data_venda,
                                    'nn_venda': ultima_venda,
                                    'cliente': ultima_cliente,
                                    'forma_de_pgto': str(forma_pgto).strip() if forma_pgto else '',
                                    'valor_venda': '',  # Vazio - pagamento adicional
                                    'entrada': entrada if entrada else 0
                                })
                                adicionais_arquivo += 1
                            except:
                                pass
            
            wb.close()
            
            if vendas_arquivo > 0 or adicionais_arquivo > 0:
                print(f'   OK: {vendas_arquivo} principais + {adicionais_arquivo} adicionais')
                total_principais += vendas_arquivo
                total_adicionais += adicionais_arquivo
        
        except Exception as e:
            print(f'   ERRO: {str(e)[:80]}')
    
    if todas_vendas:
        print(f'\n{"=" * 80}')
        print(f'SALVANDO: {loja.upper()}')
        print('=' * 80)
        
        df = pd.DataFrame(todas_vendas)
        df = df.sort_values('data_movimento')
        
        arquivo_saida = loja_dir / f'vendas_{loja}_consolidado_v2.csv'
        df.to_csv(arquivo_saida, sep=';', index=False, encoding='utf-8-sig')
        
        print(f'\nESTATISTICAS:')
        print(f'   - Total de linhas: {len(df)}')
        print(f'   - Vendas principais: {total_principais}')
        print(f'   - Pagamentos adicionais: {total_adicionais}')
        print(f'   - Vendas unicas: {df["nn_venda"].nunique()}')
        print(f'   - Periodo: {df["data_movimento"].min()} a {df["data_movimento"].max()}')
        print(f'\nArquivo salvo: {arquivo_saida.name}')

print(f'\n{"=" * 80}')
print('PROCESSAMENTO CONCLUIDO!')
print('=' * 80)
