"""
Extração completa de vendas de todos os arquivos Excel
Unifica em CSV por loja com data completa
"""

import openpyxl
import pandas as pd
from pathlib import Path
from datetime import datetime
import re

# Diretório base
base_dir = Path('dados_processados/originais/cxs/planilhas_originais')

# Lojas
lojas = ['maua', 'perus', 'rio_pequeno', 'sao_mateus', 'suzano', 'suzano2']

print('=' * 80)
print('EXTRAÇÃO COMPLETA DE VENDAS - TODOS OS ARQUIVOS EXCEL')
print('=' * 80)

def extrair_mes_ano_arquivo(nome_arquivo):
    """Extrai mês e ano do nome do arquivo (ex: jan_24.xlsx -> 01/2024)"""
    meses = {
        'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
        'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
    }
    
    # Extrair mês e ano do nome
    match = re.match(r'([a-z]{3})_(\d{2})\.xlsx', nome_arquivo.lower())
    if match:
        mes_str, ano_str = match.groups()
        mes = meses.get(mes_str, 1)
        ano = 2000 + int(ano_str)  # 24 -> 2024
        return mes, ano
    
    return None, None

def encontrar_linha_limite(ws):
    """Encontra a linha 'Restante Entrada' que marca o fim das vendas"""
    for row_idx in range(1, min(100, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value and 'restante entrada' in str(cell.value).lower():
                return row_idx
    return 50  # Default se não encontrar

def extrair_vendas_aba(ws, dia, mes, ano):
    """Extrai vendas de uma aba (dia) - incluindo múltiplas formas de pagamento"""
    vendas = []
    
    # Encontrar limite inferior
    linha_limite = encontrar_linha_limite(ws)
    
    # Variáveis para armazenar venda anterior (para pagamentos adicionais)
    ultima_venda = None
    ultima_cliente = None
    
    # Extrair vendas (linha 6 até linha_limite - 1)
    for row_idx in range(6, linha_limite):
        # Colunas: E=Nº Venda, F=Cliente, G=Forma Pgto, H=Valor Venda, I=Entrada
        num_venda = ws.cell(row_idx, 5).value  # Coluna E
        cliente = ws.cell(row_idx, 6).value    # Coluna F
        forma_pgto = ws.cell(row_idx, 7).value # Coluna G
        valor_venda = ws.cell(row_idx, 8).value # Coluna H
        entrada = ws.cell(row_idx, 9).value    # Coluna I
        
        # Criar data completa
        data_venda = f'{ano}-{mes:02d}-{dia:02d}'
        
        # CASO 1: Tem número de venda e cliente (linha principal)
        if num_venda and cliente:
            # Filtrar cabeçalhos repetidos
            if str(num_venda).strip().lower() not in ['nº venda', 'vendas', 'restante entrada']:
                try:
                    ultima_venda = str(num_venda).strip()
                    ultima_cliente = str(cliente).strip()
                    
                    vendas.append({
                        'data_movimento': data_venda,
                        'nn_venda': ultima_venda,
                        'cliente': ultima_cliente,
                        'forma_de_pgto': str(forma_pgto).strip() if forma_pgto else '',
                        'valor_venda': valor_venda if valor_venda else 0,
                        'entrada': entrada if entrada else 0
                    })
                except Exception as e:
                    print(f'      ⚠️  Erro na linha {row_idx}: {e}')
        
        # CASO 2: Não tem OS/Cliente mas tem forma_pgto ou entrada (pagamento adicional)
        elif (not num_venda or not cliente) and (forma_pgto or entrada):
            # Usar dados da venda anterior
            if ultima_venda and ultima_cliente:
                try:
                    vendas.append({
                        'data_movimento': data_venda,
                        'nn_venda': ultima_venda,
                        'cliente': ultima_cliente,
                        'forma_de_pgto': str(forma_pgto).strip() if forma_pgto else '',
                        'valor_venda': '',  # Deixar vazio - é pagamento adicional
                        'entrada': entrada if entrada else 0
                    })
                except Exception as e:
                    print(f'      ⚠️  Erro na linha adicional {row_idx}: {e}')
    
    return vendas

# Processar cada loja
for loja in lojas:
    loja_dir = base_dir / loja
    
    if not loja_dir.exists():
        continue
    
    print(f'\n{"=" * 80}')
    print(f'LOJA: {loja.upper()}')
    print('=' * 80)
    
    # Listar arquivos Excel
    arquivos = sorted(loja_dir.glob('*.xlsx'))
    print(f'\nTotal de arquivos: {len(arquivos)}')
    
    todas_vendas = []
    total_arquivos_processados = 0
    total_vendas_extraidas = 0
    
    for arquivo in arquivos:
        print(f'\nProcessando: {arquivo.name}')
        
        # Extrair mês e ano do nome do arquivo
        mes, ano = extrair_mes_ano_arquivo(arquivo.name)
        
        if not mes or not ano:
            print(f'   ERRO: Nao foi possivel extrair mes/ano do nome')
            continue
        
        print(f'   Periodo: {mes:02d}/{ano}')
        
        try:
            # Carregar workbook
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            
            # Processar cada aba de dia (01 a 31)
            vendas_arquivo = 0
            
            for dia in range(1, 32):
                aba_nome = f'{dia:02d}'
                
                if aba_nome not in wb.sheetnames:
                    continue
                
                ws = wb[aba_nome]
                vendas_dia = extrair_vendas_aba(ws, dia, mes, ano)
                
                if vendas_dia:
                    todas_vendas.extend(vendas_dia)
                    vendas_arquivo += len(vendas_dia)
            
            wb.close()
            
            if vendas_arquivo > 0:
                print(f'   ✅ Extraídas {vendas_arquivo} vendas')
                total_vendas_extraidas += vendas_arquivo
                total_arquivos_processados += 1
            else:
                print(f'   ⚠️  Nenhuma venda encontrada')
        
        except Exception as e:
            print(f'   ❌ Erro: {str(e)[:100]}')
    
    # Salvar CSV na pasta da loja
    if todas_vendas:
        print(f'\n{"=" * 80}')
        print(f'SALVANDO VENDAS DA LOJA {loja.upper()}')
        print('=' * 80)
        
        df = pd.DataFrame(todas_vendas)
        
        # Ordenar por data
        df = df.sort_values('data_movimento')
        
        # Nome do arquivo de saída
        arquivo_saida = loja_dir / f'vendas_{loja}_consolidado.csv'
        
        # Salvar
        df.to_csv(arquivo_saida, sep=';', index=False, encoding='utf-8-sig')
        
        print(f'\n📊 ESTATÍSTICAS:')
        print(f'   - Arquivos processados: {total_arquivos_processados}')
        print(f'   - Total de vendas: {len(df)}')
        print(f'   - Período: {df["data_movimento"].min()} a {df["data_movimento"].max()}')
        print(f'   - Vendas únicas: {df["nn_venda"].nunique()}')
        print(f'\n💾 Arquivo salvo: {arquivo_saida.name}')
        print(f'   📍 Local: {arquivo_saida.parent}')
    else:
        print(f'\n⚠️  Nenhuma venda encontrada na loja {loja.upper()}')

print(f'\n{"=" * 80}')
print('✅ PROCESSAMENTO CONCLUÍDO!')
print('=' * 80)
